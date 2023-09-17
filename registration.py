import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook, load_workbook

def save_to_excel():
    name = name_entry.get()
    address = address_entry.get()
    email = email_entry.get()
    contact = contact_entry.get()

    try:
        # Load existing data from the Excel file if it exists
        wb = load_workbook("data.xlsx")
        sheet = wb.active
    except FileNotFoundError:
        # Create a new workbook if the file doesn't exist
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Name", "Address", "Email", "Contact"])

    # Append the new data
    sheet.append([name, address, email, contact])

    # Save the Excel file
    wb.save("data.xlsx")

    # Clear the form fields
    name_entry.delete(0, tk.END)
    address_entry.delete(0, tk.END)
    email_entry.delete(0, tk.END)
    contact_entry.delete(0, tk.END)

# Create the main Tkinter window
root = tk.Tk()
root.title("Registration Form")

# Create a frame for the form
form_frame = ttk.Frame(root, padding=10)
form_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")

# Create labels and entry widgets for each field
name_label = ttk.Label(form_frame, text="Name:")
name_label.grid(row=0, column=0, sticky="w")
name_entry = ttk.Entry(form_frame)
name_entry.grid(row=0, column=1, sticky="e")

address_label = ttk.Label(form_frame, text="Address:")
address_label.grid(row=1, column=0, sticky="w")
address_entry = ttk.Entry(form_frame)
address_entry.grid(row=1, column=1, sticky="e")

email_label = ttk.Label(form_frame, text="Email:")
email_label.grid(row=2, column=0, sticky="w")
email_entry = ttk.Entry(form_frame)
email_entry.grid(row=2, column=1, sticky="e")

contact_label = ttk.Label(form_frame, text="Contact No:")
contact_label.grid(row=3, column=0, sticky="w")
contact_entry = ttk.Entry(form_frame)
contact_entry.grid(row=3, column=1, sticky="e")

# Create a label for the heading
heading_label = ttk.Label(root, text="Registration Form", font=("Times New Roman", 20, "bold"))
heading_label.grid(row=0, column=0, columnspan=2, padx=10, pady=(10, 0))

# Create a submit button
submit_button = ttk.Button(root, text="Submit", command=save_to_excel)
submit_button.grid(row=2, column=0, columnspan=2, pady=(10, 20))

# Configure grid weights for auto resizing
root.columnconfigure(0, weight=1)
root.rowconfigure(1, weight=1)

# Start the Tkinter main loop
root.mainloop()
